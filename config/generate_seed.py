import openpyxl
import datetime
import re

def clean_val(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s != '' else None

def clean_num(v):
    if v is None:
        return 0.0
    try:
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).replace('S/', '').replace('$', '').replace(',', '').strip()
        return float(s)
    except:
        return 0.0

def clean_date(v):
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    if re.match(r'^\d{4}-\d{2}-\d{2}', s):
        return s[:10]
    return None

wb = openpyxl.load_workbook(r'c:\xampp\htdocs\carpicenter_sys\Cuentas por Cobrar y Pagar 2026_Drive.xlsx', data_only=True)

sql_statements = []

# 1. Cta x C
ws = wb['Cta x C']
for row in list(ws.iter_rows(values_only=True))[3:]: # Skip header
    ref = clean_val(row[0])
    ft_lt = clean_val(row[1])
    cliente = clean_val(row[2])
    f_venc = clean_date(row[3])
    monto_total = clean_num(row[4])
    banco = clean_val(row[5])
    monto_pagado = clean_num(row[6])
    fecha_pago = clean_date(row[7])
    estado = clean_val(row[8]) or ('PAGADO' if monto_pagado >= monto_total and monto_total > 0 else 'PENDIENTE')
    
    if cliente or ft_lt or monto_total > 0:
        sql_statements.append(f"""
INSERT INTO finanzas_cuentas_cobrar (referencia, ft_lt, cliente, f_venc, monto_total, banco, monto_pagado, fecha_pago, estado)
VALUES ({f"'{ref}'" if ref else 'NULL'}, {f"'{ft_lt}'" if ft_lt else 'NULL'}, '{cliente or "CLIENTE VARIOS"}', {f"'{f_venc}'" if f_venc else 'NULL'}, {monto_total}, {f"'{banco}'" if banco else 'NULL'}, {monto_pagado}, {f"'{fecha_pago}'" if fecha_pago else 'NULL'}, '{estado}');
""".strip())

# 2. SUNAT
ws = wb['Sunat']
for row in list(ws.iter_rows(values_only=True))[2:]:
    nro_letra = clean_val(row[0])
    cod = clean_val(row[1])
    tributo = clean_val(row[2])
    periodo = clean_val(row[3])
    importe = clean_num(row[4])
    estado = clean_val(row[5]) or 'PENDIENTE'
    f_pago = clean_date(row[6])
    obs = clean_val(row[7])
    
    if tributo or cod or importe > 0:
        sql_statements.append(f"""
INSERT INTO finanzas_sunat (nro_letra, cod, tributo, periodo, importe, estado, f_pago, observacion)
VALUES ({f"'{nro_letra}'" if nro_letra else 'NULL'}, {f"'{cod}'" if cod else 'NULL'}, '{tributo or "IMPUESTO SUNAT"}', {f"'{periodo}'" if periodo else 'NULL'}, {importe}, '{estado}', {f"'{f_pago}'" if f_pago else 'NULL'}, {f"'{obs}'" if obs else 'NULL'});
""".strip())

# 3. SAT
ws = wb['Sat']
for row in list(ws.iter_rows(values_only=True))[2:]:
    f_emision = clean_date(row[0])
    nro_letra = clean_val(row[1])
    tipo_inf = clean_val(row[2])
    nro_doc = clean_val(row[3])
    por_pagar = clean_num(row[4])
    f_pago = clean_date(row[5])
    estado = clean_val(row[6]) or 'PENDIENTE'
    obs = clean_val(row[7])
    
    if tipo_inf or nro_doc or por_pagar > 0:
        sql_statements.append(f"""
INSERT INTO finanzas_sat (f_emision, nro_letra, tipo_infraccion, nro_documento, por_pagar, f_pago, estado, observacion)
VALUES ({f"'{f_emision}'" if f_emision else 'NULL'}, {f"'{nro_letra}'" if nro_letra else 'NULL'}, '{tipo_inf or "TRIBUTO MUNICIPAL/PAPELETA"}', {f"'{nro_doc}'" if nro_doc else 'NULL'}, {por_pagar}, {f"'{f_pago}'" if f_pago else 'NULL'}, '{estado}', {f"'{obs}'" if obs else 'NULL'});
""".strip())

# 4. Permisos
ws = wb['Permisos']
for row in list(ws.iter_rows(values_only=True))[1:]:
    tienda = clean_val(row[2])
    direccion = clean_val(row[1])
    f_servicio = clean_date(row[3])
    f_venc = clean_date(row[4])
    obs = clean_val(row[6])
    
    if direccion or tienda:
        sql_statements.append(f"""
INSERT INTO finanzas_permisos (titulo, direccion_tienda, tienda, f_servicio, f_venc, observacion)
VALUES ('CERTIFICADO ITSE / DEFENSA CIVIL', {f"'{direccion}'" if direccion else 'NULL'}, {f"'{tienda}'" if tienda else 'NULL'}, {f"'{f_servicio}'" if f_servicio else 'NULL'}, {f"'{f_venc}'" if f_venc else 'NULL'}, {f"'{obs}'" if obs else 'NULL'});
""".strip())

# Write SQL file
with open(r'c:\xampp\htdocs\carpicenter_sys\config\seed_finanzas.sql', 'w', encoding='utf-8') as f:
    f.write("\n".join(sql_statements))

print(f"Generated {len(sql_statements)} SQL statements.")
